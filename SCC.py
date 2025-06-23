import tkinter as tk
from tkinter import filedialog, messagebox
import os
from PIL import Image, ImageStat
import openpyxl
import numpy as np # Import numpy for statistical calculations

# Main application class for the Image Color Extractor
class ImageColorExtractorApp:
    def __init__(self, master):
        self.master = master
        master.title("Image Color Extractor") # Set the window title
        master.geometry("600x300") # Set initial window size
        master.resizable(False, False) # Make the window non-resizable

        # Styling for a cleaner look
        self.bg_color = "#f0f0f0"
        self.button_color = "#4CAF50" # Green for buttons
        self.button_fg = "white"
        self.label_fg = "#333333"
        self.font_main = ("Arial", 12)
        self.font_button = ("Arial", 12, "bold")

        master.configure(bg=self.bg_color) # Set background color for the main window

        # Variable to store the selected folder path
        self.folder_path = tk.StringVar(master)
        self.folder_path.set("No folder selected")

        # Variable to display status messages to the user
        self.status_message = tk.StringVar(master)
        self.status_message.set("Ready")

        # --- GUI Elements ---

        # Frame for folder selection
        self.folder_frame = tk.Frame(master, bg=self.bg_color, pady=10)
        self.folder_frame.pack(pady=10)

        # Label to display the selected folder path
        self.folder_label = tk.Label(self.folder_frame, textvariable=self.folder_path,
                                     bg=self.bg_color, fg=self.label_fg, font=self.font_main,
                                     width=50, anchor="w", relief="sunken", bd=1)
        self.folder_label.pack(side=tk.LEFT, padx=5)

        # Button to browse for a folder
        self.browse_button = tk.Button(self.folder_frame, text="Browse Folder",
                                       command=self.browse_folder,
                                       bg=self.button_color, fg=self.button_fg,
                                       font=self.font_button, relief="raised", bd=2,
                                       cursor="hand2")
        self.browse_button.pack(side=tk.LEFT, padx=5)

        # Frame for action button
        self.action_frame = tk.Frame(master, bg=self.bg_color, pady=10)
        self.action_frame.pack(pady=10)

        # Button to start processing images
        self.process_button = tk.Button(self.action_frame, text="Process Images",
                                        command=self.process_images,
                                        bg="#007BFF", fg=self.button_fg, # Blue for process button
                                        font=self.font_button, relief="raised", bd=2,
                                        cursor="hand2")
        self.process_button.pack(padx=5)

        # Label to display status messages
        self.status_label = tk.Label(master, textvariable=self.status_message,
                                     bg=self.bg_color, fg="#FF4500", # OrangeRed for status
                                     font=("Arial", 10, "italic"))
        self.status_label.pack(pady=10)

    # Function to open a file dialog and select a folder
    def browse_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.folder_path.set(folder_selected)
            self.status_message.set(f"Folder selected: {folder_selected}")
        else:
            self.status_message.set("Folder selection cancelled.")

    # Function to get various RGB statistics from a middle rectangular region of an image
    def get_middle_rectangle_rgb_stats(self, image_path, percentage=0.15):
        try:
            with Image.open(image_path) as img:
                # Ensure image is in RGB format
                if img.mode != 'RGB':
                    img = img.convert('RGB')

                width, height = img.size

                # Calculate the dimensions of the middle rectangle
                rect_width = int(width * percentage)
                rect_height = int(height * percentage)

                # Calculate top-left corner coordinates for the centered rectangle
                left = (width - rect_width) // 2
                top = (height - rect_height) // 2
                right = left + rect_width
                bottom = top + rect_height

                # Ensure dimensions are at least 1x1 to avoid errors with very small images
                if rect_width < 1: rect_width = 1
                if rect_height < 1: rect_height = 1

                # Correct bounding box if it goes out of bounds for very small images or edge cases
                left = max(0, left)
                top = max(0, top)
                right = min(width, right)
                bottom = min(height, bottom)

                # Crop the image to the defined rectangle
                cropped_img = img.crop((left, top, right, bottom))

                # Get pixel data as a list of (R, G, B) tuples
                pixel_data = np.array(cropped_img.getdata())

                # If grayscale, convert to 3 channels for consistent processing
                if pixel_data.ndim == 1:
                    pixel_data = np.stack([pixel_data, pixel_data, pixel_data], axis=-1)
                elif pixel_data.shape[1] == 4: # Handle RGBA, discard A
                    pixel_data = pixel_data[:, :3]

                # Calculate statistics for each channel (R, G, B)
                mean_rgb = np.mean(pixel_data, axis=0).astype(int)
                median_rgb = np.median(pixel_data, axis=0).astype(int)
                min_rgb = np.min(pixel_data, axis=0).astype(int)
                max_rgb = np.max(pixel_data, axis=0).astype(int)

                # Calculate quartiles
                q1_rgb = np.percentile(pixel_data, 25, axis=0).astype(int)
                q3_rgb = np.percentile(pixel_data, 75, axis=0).astype(int)

                # Calculate mean of interquartile range (IQR) values
                # Filter pixels within the IQR for each channel
                iqr_mean_rgb = []
                for i in range(3): # For R, G, B channels
                    channel_values = pixel_data[:, i]
                    lower_bound = q1_rgb[i]
                    upper_bound = q3_rgb[i]

                    iqr_values = channel_values[(channel_values >= lower_bound) & (channel_values <= upper_bound)]
                    if iqr_values.size > 0:
                        iqr_mean_rgb.append(int(np.mean(iqr_values)))
                    else:
                        iqr_mean_rgb.append(int(mean_rgb[i])) # Fallback to mean if no values in IQR (unlikely but safe)

                iqr_mean_rgb = np.array(iqr_mean_rgb).astype(int)

                return {
                    "mean": mean_rgb.tolist(),
                    "median": median_rgb.tolist(),
                    "min": min_rgb.tolist(),
                    "max": max_rgb.tolist(),
                    "q1": q1_rgb.tolist(),
                    "q3": q3_rgb.tolist(),
                    "iqr_mean": iqr_mean_rgb.tolist()
                }, cropped_img # Return both stats dictionary and the cropped image
        except Exception as e:
            # Handle errors during image processing (e.g., corrupted file)
            print(f"Error processing image {image_path}: {e}")
            return None, None # Return None for both in case of error

    # Function to process all JPG images in the selected folder and save data to Excel
    def process_images(self):
        folder = self.folder_path.get()
        if not os.path.isdir(folder):
            messagebox.showerror("Error", "Please select a valid folder first.")
            self.status_message.set("Error: No valid folder selected.")
            return

        self.status_message.set("Processing images, please wait...")
        self.master.update_idletasks() # Update GUI to show message

        cropped_output_dir = os.path.join(folder, "cropped_images")
        os.makedirs(cropped_output_dir, exist_ok=True) # Create output directory for cropped images

        # Create a new Excel workbook
        wb = openpyxl.Workbook()
        # The first sheet is for the summary data
        ws_summary = wb.active
        ws_summary.title = "Image Colors Summary"

        # Write header row for summary sheet with all new statistics
        ws_summary.append([
            "Filename",
            "Mean R", "Mean G", "Mean B",
            "Median R", "Median G", "Median B",
            "Min R", "Min G", "Min B",
            "Max R", "Max G", "Max B",
            "Q1 R", "Q1 G", "Q1 B",
            "Q3 R", "Q3 G", "Q3 B",
            "IQR Mean R", "IQR Mean G", "IQR Mean B"
        ])


        for filename in os.listdir(folder):
            if filename.lower().endswith(('.jpg', '.jpeg')):
                image_path = os.path.join(folder, filename)
                stats, cropped_img = self.get_middle_rectangle_rgb_stats(image_path) # Get stats dictionary and cropped image

                if stats and cropped_img:
                    # Append summary data to the first sheet
                    row_data = [
                        filename,
                        *stats["mean"],
                        *stats["median"],
                        *stats["min"],
                        *stats["max"],
                        *stats["q1"],
                        *stats["q3"],
                        *stats["iqr_mean"]
                    ]
                    ws_summary.append(row_data)

                    # Save the cropped image
                    base, ext = os.path.splitext(filename)
                    cropped_filename = f"{base}_cropped{ext}"
                    cropped_image_path = os.path.join(cropped_output_dir, cropped_filename)
                    try:
                        cropped_img.save(cropped_image_path)
                    except Exception as e:
                        print(f"Error saving cropped image {cropped_filename}: {e}")
                        self.status_message.set(f"Error saving cropped {cropped_filename}")
                        self.master.update_idletasks()
                else:
                    self.status_message.set(f"Skipped: {filename} (Error processing)")
                    self.master.update_idletasks()

        if not ws_summary.max_row > 1: # Check if only header is present
            messagebox.showinfo("No Images Found", "No JPG files found in the selected folder, or none could be processed.")
            self.status_message.set("No JPG images found or processed.")
            # Remove the empty summary sheet if no images were processed
            wb.remove(ws_summary)
            # If no other sheets were created, create a dummy one to avoid error on save
            if not wb.sheetnames:
                wb.create_sheet("Empty")

            return

        # Define the output Excel file path
        output_excel_path = os.path.join(folder, "image_colors.xlsx")

        try:
            # Save the workbook
            wb.save(output_excel_path)
            messagebox.showinfo("Success", f"Excel file created successfully at:\n{output_excel_path}\nCropped images saved to:\n{cropped_output_dir}")
            self.status_message.set(f"Processing complete! Results saved to {output_excel_path} and cropped images to {cropped_output_dir}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel file: {e}")
            self.status_message.set(f"Error saving Excel: {e}")

# Main entry point for the application
if __name__ == "__main__":
    # Create the main Tkinter window
    root = tk.Tk()
    # Instantiate the application class
    app = ImageColorExtractorApp(root)
    # Start the Tkinter event loop
    root.mainloop()
